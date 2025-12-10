import csv
import json

import openpyxl
from zipfile import BadZipFile

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import user_passes_test
from django.contrib.sessions.models import Session
from django.db.models import Q
from django.forms import modelformset_factory
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date

from .forms import (
    EmployeeImportForm,
    HREmployeeCreateForm,
    HREmployeeUpdateForm,
    LeaveBalanceForm,
)
from .models import Department, EmployeeProfile, LeaveBalance, LeaveRequest, LeaveType
from .services import create_default_leave_balances

User = get_user_model()


def is_hr(user):
    return (
        user.is_superuser
        or user.is_staff
        or user.groups.filter(name="HR").exists()
    )


@user_passes_test(is_hr)
def hr_leave_dashboard(request):
    qs = _get_filtered_leaves(request)

    status = request.GET.get("status") or ""
    department_id = request.GET.get("department") or ""
    leave_type_id = request.GET.get("leave_type") or ""
    employee_id = request.GET.get("employee") or ""
    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""

    context = {
        "leaves": qs,
        "statuses": LeaveRequest.STATUS_CHOICES,
        "departments": Department.objects.all(),
        "leave_types": LeaveType.objects.all(),
        "employees": EmployeeProfile.objects.select_related("user").all(),
        "filter_status": status,
        "filter_department": department_id,
        "filter_leave_type": leave_type_id,
        "filter_employee": employee_id,
        "filter_date_from": date_from,
        "filter_date_to": date_to,
    }
    return render(request, "leave_app/hr/hr_leave_dashboard.html", context)


@user_passes_test(is_hr)
def hr_employee_create(request):
    if request.method == "POST":
        form = HREmployeeCreateForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data

            user = User.objects.create_user(
                username=cd["username"],
                password=cd["password"],
            )

            profile = EmployeeProfile.objects.create(
                user=user,
                employee_code=cd["employee_code"],
                department=cd["department"],
                manager=cd["manager"],
            )

            create_default_leave_balances(profile)

            messages.success(request, f"สร้างพนักงาน {profile} เรียบร้อยแล้ว")
            return redirect("leave_app:hr_employee_list")
    else:
        form = HREmployeeCreateForm()

    return render(request, "leave_app/hr/hr_employee_create.html", {"form": form})


@user_passes_test(is_hr)
def hr_employee_list(request):
    q = request.GET.get("q", "").strip()
    department_id = request.GET.get("department") or ""
    status = request.GET.get("status") or "active"  # active / inactive / all

    employees = EmployeeProfile.objects.select_related(
        "user", "department", "manager"
    )

    if q:
        employees = employees.filter(
            Q(employee_code__icontains=q)
            | Q(user__username__icontains=q)
            | Q(user__first_name__icontains=q)
            | Q(user__last_name__icontains=q)
        )

    if department_id:
        employees = employees.filter(department_id=department_id)

    if status == "active":
        employees = employees.filter(user__is_active=True)
    elif status == "inactive":
        employees = employees.filter(user__is_active=False)

    employees = employees.order_by("employee_code")

    context = {
        "employees": employees,
        "departments": Department.objects.all(),
        "q": q,
        "filter_department": department_id,
        "filter_status": status,
    }
    return render(request, "leave_app/hr/hr_employee_list.html", context)


@user_passes_test(is_hr)
def hr_employee_edit(request, pk):
    profile = get_object_or_404(
        EmployeeProfile.objects.select_related("user", "department", "manager"),
        pk=pk,
    )
    user_obj = profile.user

    if request.method == "POST":
        form = HREmployeeUpdateForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "อัปเดตข้อมูลพนักงานเรียบร้อยแล้ว")
            if "stay" in request.POST:
                return redirect("leave_app:hr_employee_edit", pk=pk)
            return redirect("leave_app:hr_employee_list")
    else:
        form = HREmployeeUpdateForm(instance=profile)

    context = {
        "profile": profile,
        "user_obj": user_obj,
        "form": form,
    }
    return render(request, "leave_app/hr/hr_employee_edit.html", context)


@user_passes_test(is_hr)
def hr_employee_toggle_active(request, pk):
    profile = get_object_or_404(EmployeeProfile.objects.select_related("user"), pk=pk)
    if request.method == "POST":
        user = profile.user
        user.is_active = not user.is_active
        user.save()
        
        # ถ้าปิดการใช้งาน → เคลียร์ทุก session ของ user นี้
        if not user.is_active:
            sessions = Session.objects.filter(expire_date__gte=timezone.now())
            for s in sessions:
                data = s.get_decoded()
                if data.get("_auth_user_id") == str(user.id):
                    s.delete()
        
        status = "เปิดใช้งาน" if user.is_active else "ปิดการใช้งาน"
        messages.success(request, f"{status}บัญชีผู้ใช้เรียบร้อยแล้ว")
    return redirect("leave_app:hr_employee_edit", pk=pk)


@user_passes_test(is_hr)
def hr_employee_import(request):
    if request.method == "POST":
        form = EmployeeImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data["file"]
            try:
                wb = openpyxl.load_workbook(file)
            except BadZipFile:
                messages.error(
                    request,
                    "ไฟล์นี้ไม่ใช่ Excel .xlsx กรุณาส่งออกเป็น .xlsx แล้วลองใหม่อีกครั้ง",
                )
                return redirect("leave_app:hr_employee_import")

            ws = wb.active
            created_count = 0

            for idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                username, password, employee_code, dept_code, manager_username = row[:5]
                if not username:
                    continue

                user, user_created = User.objects.get_or_create(username=username)
                if user_created:
                    user.set_password(password or username)
                    user.save()

                department = None
                if dept_code:
                    department, _ = Department.objects.get_or_create(
                        code=str(dept_code),
                        defaults={"name": str(dept_code)},
                    )

                manager = None
                if manager_username:
                    manager = User.objects.filter(
                        username=str(manager_username)
                    ).first()

                profile, prof_created = EmployeeProfile.objects.get_or_create(
                    user=user,
                    defaults={
                        "employee_code": employee_code or f"EMP{user.id:04d}",
                        "department": department,
                        "manager": manager,
                    },
                )

                create_default_leave_balances(profile)
                created_count += 1

            messages.success(request, f"นำเข้าพนักงานสำเร็จ {created_count} รายการ")
            return redirect("leave_app:hr_employee_list")
    else:
        form = EmployeeImportForm()

    return render(request, "leave_app/hr/hr_employee_import.html", {"form": form})


@user_passes_test(is_hr)
def hr_leave_balance_manage(request):
    employees = EmployeeProfile.objects.select_related("user").all()

    year_param = request.GET.get("year")
    employee_id = request.GET.get("employee")

    try:
        year = int(year_param) if year_param else timezone.now().year
    except ValueError:
        year = timezone.now().year

    selected_employee = None
    formset = None

    if employee_id:
        selected_employee = get_object_or_404(EmployeeProfile, pk=employee_id)

        create_default_leave_balances(selected_employee, year)

        qs = LeaveBalance.objects.filter(
            employee=selected_employee, year=year
        ).select_related("leave_type")

        LeaveBalanceFormSet = modelformset_factory(
            LeaveBalance, form=LeaveBalanceForm, extra=0
        )

        if request.method == "POST":
            formset = LeaveBalanceFormSet(request.POST, queryset=qs)
            if formset.is_valid():
                formset.save()
                messages.success(request, "อัปเดตโควต้าวันลาเรียบร้อยแล้ว")
                url = reverse("leave_app:hr_leave_balance_manage")
                return redirect(f"{url}?employee={employee_id}&year={year}")
            else:
                messages.error(request, "กรุณาตรวจสอบข้อมูลโควต้าวันลาอีกครั้ง")
        else:
            formset = LeaveBalanceFormSet(queryset=qs)

    context = {
        "employees": employees,
        "selected_employee": selected_employee,
        "year": year,
        "formset": formset,
        "employee_id": employee_id or "",
    }
    return render(request, "leave_app/hr/hr_leave_balance_manage.html", context)


def _get_filtered_leaves(request):
    qs = (
        LeaveRequest.objects.select_related(
            "employee__user",
            "employee__department",
            "leave_type",
        )
        .order_by("-created_at")
    )

    status = request.GET.get("status") or ""
    department_id = request.GET.get("department") or ""
    leave_type_id = request.GET.get("leave_type") or ""
    employee_id = request.GET.get("employee") or ""
    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""

    if status:
        qs = qs.filter(status=status)
    if department_id:
        qs = qs.filter(employee__department_id=department_id)
    if leave_type_id:
        qs = qs.filter(leave_type_id=leave_type_id)
    if employee_id:
        qs = qs.filter(employee_id=employee_id)
    if date_from:
        qs = qs.filter(start_date__gte=parse_date(date_from))
    if date_to:
        qs = qs.filter(end_date__lte=parse_date(date_to))

    return qs


@user_passes_test(is_hr)
def hr_export_leaves_csv(request):
    qs = _get_filtered_leaves(request)

    response = HttpResponse(content_type="text/csv")
    filename = f"leave_requests_{timezone.now().date()}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(
        [
            "Created at",
            "Employee code",
            "Employee name",
            "Department",
            "Leave type",
            "Start date",
            "End date",
            "Half day",
            "Status",
            "Reason",
        ]
    )

    for leave in qs:
        emp = leave.employee
        user = emp.user
        writer.writerow(
            [
                leave.created_at.strftime("%Y-%m-%d %H:%M"),
                emp.employee_code,
                user.get_full_name() or user.username,
                emp.department.name if emp.department else "",
                leave.leave_type.name,
                leave.start_date,
                leave.end_date,
                "Yes" if leave.half_day else "No",
                leave.get_status_display(),
                leave.reason.replace("\n", " "),
            ]
        )

    return response


@user_passes_test(is_hr)
def hr_export_leaves_excel(request):
    qs = _get_filtered_leaves(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leave Requests"

    headers = [
        "Created at",
        "Employee code",
        "Employee name",
        "Department",
        "Leave type",
        "Start date",
        "End date",
        "Half day",
        "Status",
        "Reason",
    ]
    ws.append(headers)

    for leave in qs:
        emp = leave.employee
        user = emp.user
        ws.append(
            [
                leave.created_at.strftime("%Y-%m-%d %H:%M"),
                emp.employee_code,
                user.get_full_name() or user.username,
                emp.department.name if emp.department else "",
                leave.leave_type.name,
                leave.start_date.isoformat(),
                leave.end_date.isoformat(),
                "Yes" if leave.half_day else "No",
                leave.get_status_display(),
                leave.reason,
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"leave_requests_{timezone.now().date()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
