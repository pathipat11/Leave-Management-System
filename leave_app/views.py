from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.core.exceptions import ValidationError
from django.http import HttpResponseForbidden, HttpResponse
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login as auth_login, logout as auth_logout, get_user_model
from django.forms import modelformset_factory
from django.urls import reverse
from django.db.models import Count, Q
from django.db.models.functions import TruncMonth
from zipfile import BadZipFile

from .services import (
    approve_leave_request, reject_leave_request,
    create_default_leave_balances, notify_leave_submitted,
)
from .models import EmployeeProfile, LeaveRequest, Department, LeaveType, LeaveBalance
from django.utils.dateparse import parse_date
from .forms import (
    LeaveRequestForm,
    HREmployeeCreateForm,
    HREmployeeUpdateForm,
    EmployeeImportForm,
    LeaveBalanceForm,
)
import openpyxl, csv, json


def register(request):
    if request.method == "POST":
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()

            employee_code = f"EMP{user.id:04d}"
            profile = EmployeeProfile.objects.create(
                user=user,
                employee_code=employee_code,
                department=None,
            )

            create_default_leave_balances(profile)

            auth_login(request, user)
            messages.success(request, "สมัครสมาชิกสำเร็จแล้ว")
            return redirect("leave_app:dashboard")
    else:
        form = UserCreationForm()

    return render(request, "registration/register.html", {"form": form})


@login_required
def dashboard(request):
    profile = get_object_or_404(EmployeeProfile, user=request.user)
    recent_leaves = LeaveRequest.objects.filter(employee=profile)[:5]

    current_year = timezone.now().year
    balances = LeaveBalance.objects.filter(
        employee=profile,
        year=current_year,
    ).select_related("leave_type")

    context = {
        "profile": profile,
        "recent_leaves": recent_leaves,
        "balances": balances,
        "current_year": current_year,
    }
    return render(request, "leave_app/dashboard.html", context)



@login_required
def leave_request_list(request):
    profile = get_object_or_404(EmployeeProfile, user=request.user)
    leaves = LeaveRequest.objects.filter(employee=profile)
    return render(request, "leave_app/leave_request_list.html", {"leaves": leaves})


@login_required
def leave_request_create(request):
    profile = get_object_or_404(EmployeeProfile, user=request.user)

    if request.method == "POST":
        form = LeaveRequestForm(request.POST, request.FILES, employee_profile=profile)
        if form.is_valid():
            leave_req = form.save(commit=False)
            leave_req.employee = profile
            leave_req.save()

            notify_leave_submitted(leave_req)

            messages.success(request, "ส่งคำขอลางานเรียบร้อยแล้ว")
            return redirect("leave_app:leave_request_list")
    else:
        form = LeaveRequestForm(employee_profile=profile)

    return render(request, "leave_app/leave_request_form.html", {"form": form})


def logout_view(request):
    auth_logout(request)
    return redirect("login")


@login_required
def leave_request_cancel(request, pk):
    profile = get_object_or_404(EmployeeProfile, user=request.user)
    leave_req = get_object_or_404(LeaveRequest, pk=pk, employee=profile)

    if leave_req.status != LeaveRequest.STATUS_PENDING:
        messages.error(request, "ยกเลิกได้เฉพาะคำขอที่ยัง Pending เท่านั้น")
        return redirect("leave_app:leave_request_list")

    leave_req.status = LeaveRequest.STATUS_CANCELLED
    leave_req.cancelled_at = timezone.now()
    leave_req.save()
    messages.success(request, "ยกเลิกคำขอลาเรียบร้อยแล้ว")
    return redirect("leave_app:leave_request_list")


def is_manager(user):
    return user.is_superuser or user.groups.filter(name="MANAGER").exists()


def is_hr(user):
    return user.is_superuser or user.is_staff or user.groups.filter(name="HR").exists()


def is_ceo(user):
    return user.is_superuser or user.groups.filter(name="CEO").exists()


@user_passes_test(is_manager)
def manager_leave_list(request):
    subordinates = EmployeeProfile.objects.filter(manager=request.user)

    # คำขอที่ยัง Pending (รออนุมัติ)
    pending_leaves = LeaveRequest.objects.filter(
        employee__in=subordinates,
        status=LeaveRequest.STATUS_PENDING,
    ).select_related("employee", "leave_type")

    # ประวัติคำขอลา (ทุกสถานะ ยกเว้น Pending)
    history_leaves = LeaveRequest.objects.filter(
        employee__in=subordinates,
    ).exclude(
        status=LeaveRequest.STATUS_PENDING,
    ).select_related(
        "employee", "leave_type", "approver"
    ).order_by("-updated_at")

    context = {
        "pending_leaves": pending_leaves,
        "history_leaves": history_leaves,
    }
    return render(request, "leave_app/manager_leave_list.html", context)



@user_passes_test(is_manager)
def manager_leave_detail(request, pk):
    leave_req = get_object_or_404(
        LeaveRequest.objects.select_related("employee", "leave_type"),
        pk=pk,
    )

    if leave_req.employee.manager != request.user and not request.user.is_superuser:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์ดูคำขอนี้")

    if request.method == "POST":
        action = request.POST.get("action")
        comment = request.POST.get("comment", "")

        try:
            if action == "approve":
                approve_leave_request(leave_req, approver=request.user, comment=comment)
                messages.success(request, "อนุมัติคำขอลาเรียบร้อยแล้ว")
            elif action == "reject":
                reject_leave_request(leave_req, approver=request.user, comment=comment)
                messages.success(request, "ปฏิเสธคำขอลาเรียบร้อยแล้ว")
            else:
                messages.error(request, "คำสั่งไม่ถูกต้อง")
        except ValidationError as e:
            messages.error(request, e.message)

        return redirect("leave_app:manager_leave_list")

    context = {
        "leave": leave_req,
    }
    return render(request, "leave_app/manager_leave_detail.html", context)


@user_passes_test(is_hr)
def hr_leave_dashboard(request):
    qs = _get_filtered_leaves(request)

    status = request.GET.get("status") or ""
    department_id = request.GET.get("department") or ""
    leave_type_id = request.GET.get("leave_type") or ""
    employee_id = request.GET.get("employee") or ""
    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""

    context = {
        "leaves": qs,
        "statuses": LeaveRequest.STATUS_CHOICES,
        "departments": Department.objects.all(),
        "leave_types": LeaveType.objects.all(),
        "employees": EmployeeProfile.objects.select_related("user").all(),
        "filter_status": status,
        "filter_department": department_id,
        "filter_leave_type": leave_type_id,
        "filter_employee": employee_id,
        "filter_date_from": date_from,
        "filter_date_to": date_to,
    }
    return render(request, "leave_app/hr_leave_dashboard.html", context)


User = get_user_model()


@user_passes_test(is_hr)
def hr_employee_create(request):
    """
    HR เพิ่มพนักงาน 1 คน:
    - สร้าง User
    - สร้าง EmployeeProfile
    - สร้าง LeaveBalance ให้ปีปัจจุบัน
    """
    if request.method == "POST":
        form = HREmployeeCreateForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data

            user = User.objects.create_user(
                username=cd["username"],
                password=cd["password"],
            )

            profile = EmployeeProfile.objects.create(
                user=user,
                employee_code=cd["employee_code"],
                department=cd["department"],
                manager=cd["manager"],
            )

            create_default_leave_balances(profile)

            messages.success(request, f"สร้างพนักงาน {profile} เรียบร้อยแล้ว")
            return redirect("leave_app:hr_employee_list")
    else:
        form = HREmployeeCreateForm()

    return render(request, "leave_app/hr_employee_create.html", {"form": form})


@user_passes_test(is_hr)
def hr_employee_list(request):
    """
    HR ดูรายชื่อพนักงานทั้งหมด + filter + search
    """
    q = request.GET.get("q", "").strip()
    department_id = request.GET.get("department") or ""
    status = request.GET.get("status") or "active"  # active / inactive / all

    employees = EmployeeProfile.objects.select_related("user", "department", "manager")

    if q:
        employees = employees.filter(
            Q(employee_code__icontains=q)
            | Q(user__username__icontains=q)
            | Q(user__first_name__icontains=q)
            | Q(user__last_name__icontains=q)
        )

    if department_id:
        employees = employees.filter(department_id=department_id)

    if status == "active":
        employees = employees.filter(user__is_active=True)
    elif status == "inactive":
        employees = employees.filter(user__is_active=False)

    employees = employees.order_by("employee_code")

    context = {
        "employees": employees,
        "departments": Department.objects.all(),
        "q": q,
        "filter_department": department_id,
        "filter_status": status,
    }
    return render(request, "leave_app/hr_employee_list.html", context)


@user_passes_test(is_hr)
def hr_employee_edit(request, pk):
    """
    HR แก้ไขข้อมูลพนักงาน + ข้อมูล User (ชื่อ, อีเมล, active)
    """
    profile = get_object_or_404(
        EmployeeProfile.objects.select_related("user", "department", "manager"),
        pk=pk,
    )
    user_obj = profile.user

    if request.method == "POST":
        form = HREmployeeUpdateForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "อัปเดตข้อมูลพนักงานเรียบร้อยแล้ว")
            if "stay" in request.POST:
                return redirect("leave_app:hr_employee_edit", pk=pk)
            return redirect("leave_app:hr_employee_list")
    else:
        form = HREmployeeUpdateForm(instance=profile)

    context = {
        "profile": profile,
        "user_obj": user_obj,
        "form": form,
    }
    return render(request, "leave_app/hr_employee_edit.html", context)


@user_passes_test(is_hr)
def hr_employee_toggle_active(request, pk):
    """
    HR เปิด/ปิดการใช้งาน User
    """
    profile = get_object_or_404(EmployeeProfile.objects.select_related("user"), pk=pk)
    if request.method == "POST":
        user = profile.user
        user.is_active = not user.is_active
        user.save()
        status = "เปิดใช้งาน" if user.is_active else "ปิดการใช้งาน"
        messages.success(request, f"{status}บัญชีผู้ใช้เรียบร้อยแล้ว")
    return redirect("leave_app:hr_employee_edit", pk=pk)


@user_passes_test(is_hr)
def hr_employee_import(request):
    if request.method == "POST":
        form = EmployeeImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data["file"]
            try:
                wb = openpyxl.load_workbook(file)
            except BadZipFile:
                messages.error(
                    request,
                    "ไฟล์นี้ไม่ใช่ Excel .xlsx กรุณาส่งออกเป็น .xlsx แล้วลองใหม่อีกครั้ง"
                )
                return redirect("leave_app:hr_employee_import")

            ws = wb.active
            created_count = 0

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                username, password, employee_code, dept_code, manager_username = row[:5]
                if not username:
                    continue

                user, user_created = User.objects.get_or_create(username=username)
                if user_created:
                    user.set_password(password or username)
                    user.save()

                department = None
                if dept_code:
                    department, _ = Department.objects.get_or_create(
                        code=str(dept_code),
                        defaults={"name": str(dept_code)},
                    )

                manager = None
                if manager_username:
                    manager = User.objects.filter(username=str(manager_username)).first()

                profile, prof_created = EmployeeProfile.objects.get_or_create(
                    user=user,
                    defaults={
                        "employee_code": employee_code or f"EMP{user.id:04d}",
                        "department": department,
                        "manager": manager,
                    },
                )

                create_default_leave_balances(profile)
                created_count += 1

            messages.success(request, f"นำเข้าพนักงานสำเร็จ {created_count} รายการ")
            return redirect("leave_app:hr_employee_list")
    else:
        form = EmployeeImportForm()

    return render(request, "leave_app/hr_employee_import.html", {"form": form})


@user_passes_test(is_hr)
def hr_leave_balance_manage(request):
    """
    HR จัดการ LeaveBalance:
    - เลือกพนักงาน + ปี
    - แก้ไข allocated/used ของทุก LeaveType
    """
    employees = EmployeeProfile.objects.select_related("user").all()

    year_param = request.GET.get("year")
    employee_id = request.GET.get("employee")

    try:
        year = int(year_param) if year_param else timezone.now().year
    except ValueError:
        year = timezone.now().year

    selected_employee = None
    formset = None

    if employee_id:
        selected_employee = get_object_or_404(EmployeeProfile, pk=employee_id)

        create_default_leave_balances(selected_employee, year)

        qs = LeaveBalance.objects.filter(
            employee=selected_employee,
            year=year,
        ).select_related("leave_type")

        LeaveBalanceFormSet = modelformset_factory(
            LeaveBalance, form=LeaveBalanceForm, extra=0
        )

        if request.method == "POST":
            formset = LeaveBalanceFormSet(request.POST, queryset=qs)
            if formset.is_valid():
                formset.save()
                messages.success(request, "อัปเดตโควต้าวันลาเรียบร้อยแล้ว")
                url = reverse("leave_app:hr_leave_balance_manage")
                return redirect(f"{url}?employee={employee_id}&year={year}")
            else:
                messages.error(request, "กรุณาตรวจสอบข้อมูลโควต้าวันลาอีกครั้ง")
        else:
            formset = LeaveBalanceFormSet(queryset=qs)


    context = {
        "employees": employees,
        "selected_employee": selected_employee,
        "year": year,
        "formset": formset,
        "employee_id": employee_id or "",
    }
    return render(request, "leave_app/hr_leave_balance_manage.html", context)


def _get_filtered_leaves(request):
    qs = LeaveRequest.objects.select_related(
        "employee__user",
        "employee__department",
        "leave_type",
    ).order_by("-created_at")

    status = request.GET.get("status") or ""
    department_id = request.GET.get("department") or ""
    leave_type_id = request.GET.get("leave_type") or ""
    employee_id = request.GET.get("employee") or ""
    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""

    if status:
        qs = qs.filter(status=status)
    if department_id:
        qs = qs.filter(employee__department_id=department_id)
    if leave_type_id:
        qs = qs.filter(leave_type_id=leave_type_id)
    if employee_id:
        qs = qs.filter(employee_id=employee_id)
    if date_from:
        qs = qs.filter(start_date__gte=parse_date(date_from))
    if date_to:
        qs = qs.filter(end_date__lte=parse_date(date_to))

    return qs


@user_passes_test(is_hr)
def hr_export_leaves_csv(request):
    qs = _get_filtered_leaves(request)

    response = HttpResponse(content_type="text/csv")
    filename = f"leave_requests_{timezone.now().date()}.csv"
    response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'

    writer = csv.writer(response)
    writer.writerow([
        "Created at", "Employee code", "Employee name", "Department",
        "Leave type", "Start date", "End date", "Half day",
        "Status", "Reason",
    ])

    for leave in qs:
        emp = leave.employee
        user = emp.user
        writer.writerow([
            leave.created_at.strftime("%Y-%m-%d %H:%M"),
            emp.employee_code,
            user.get_full_name() or user.username,
            emp.department.name if emp.department else "",
            leave.leave_type.name,
            leave.start_date,
            leave.end_date,
            "Yes" if leave.half_day else "No",
            leave.get_status_display(),
            leave.reason.replace("\n", " "),
        ])

    return response


@user_passes_test(is_hr)
def hr_export_leaves_excel(request):
    qs = _get_filtered_leaves(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leave Requests"

    headers = [
        "Created at", "Employee code", "Employee name", "Department",
        "Leave type", "Start date", "End date", "Half day",
        "Status", "Reason",
    ]
    ws.append(headers)

    for leave in qs:
        emp = leave.employee
        user = emp.user
        ws.append([
            leave.created_at.strftime("%Y-%m-%d %H:%M"),
            emp.employee_code,
            user.get_full_name() or user.username,
            emp.department.name if emp.department else "",
            leave.leave_type.name,
            leave.start_date.isoformat(),
            leave.end_date.isoformat(),
            "Yes" if leave.half_day else "No",
            leave.get_status_display(),
            leave.reason,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"leave_requests_{timezone.now().date()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    wb.save(response)
    return response


@user_passes_test(is_ceo)
def ceo_dashboard(request):
    year_param = request.GET.get("year")
    try:
        year = int(year_param) if year_param else timezone.now().year
    except ValueError:
        year = timezone.now().year

    total_employees = EmployeeProfile.objects.filter(user__is_active=True).count()

    qs_year = LeaveRequest.objects.filter(start_date__year=year)

    total_requests = qs_year.count()
    pending_count = qs_year.filter(status=LeaveRequest.STATUS_PENDING).count()
    approved_count = qs_year.filter(status=LeaveRequest.STATUS_APPROVED).count()
    rejected_count = qs_year.filter(status=LeaveRequest.STATUS_REJECTED).count()
    cancelled_count = qs_year.filter(status=LeaveRequest.STATUS_CANCELLED).count()

    monthly_qs = qs_year.annotate(
        month=TruncMonth("start_date")
    ).values("month").annotate(
        count=Count("id")
    ).order_by("month")

    monthly_labels = [item["month"].strftime("%b") for item in monthly_qs]
    monthly_counts = [item["count"] for item in monthly_qs]

    dept_qs = qs_year.values(
        "employee__department__name"
    ).annotate(
        count=Count("id")
    ).order_by("-count")

    department_labels = [
        (item["employee__department__name"] or "No Dept") for item in dept_qs
    ]
    department_counts = [item["count"] for item in dept_qs]

    type_qs = qs_year.values(
        "leave_type__name"
    ).annotate(
        count=Count("id")
    ).order_by("-count")

    leave_type_labels = [item["leave_type__name"] for item in type_qs]
    leave_type_counts = [item["count"] for item in type_qs]

    context = {
        "year": year,
        "total_employees": total_employees,
        "total_requests": total_requests,
        "pending_count": pending_count,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
        "cancelled_count": cancelled_count,
        "monthly_labels_json": json.dumps(monthly_labels),
        "monthly_counts_json": json.dumps(monthly_counts),
        "department_labels_json": json.dumps(department_labels),
        "department_counts_json": json.dumps(department_counts),
        "leave_type_labels_json": json.dumps(leave_type_labels),
        "leave_type_counts_json": json.dumps(leave_type_counts),
    }
    return render(request, "leave_app/ceo_dashboard.html", context)
